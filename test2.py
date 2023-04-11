import json
from openpyxl import Workbook
import Levenshtein
from baidutran import baidutrans
from langdetect import detect

# 创建一个Workbook对象
workbook = Workbook()

# 获取当前活动的Worksheet对象
worksheet = workbook.active

with open('data.json', 'r') as f:
    data = json.load(f)
with open('naiya.json', 'r') as f:
    data2 = json.load(f)


def find_most_similar(target, string_list):
    distances = [Levenshtein.distance(target, string) for string in string_list]
    min_index = distances.index(min(distances))
    return min_index, distances[min_index]


num_same = 0
num_not_same = 0

jifan_list = []

for key in data2:
    if key in data:
        if len(data[key]['context']) == len(data2[key]['context']):
            num_same += 1
        else:
            num_not_same += 1
            worksheet.cell(row=1, column=4 * num_not_same - 2, value='机器翻译')
            worksheet.cell(row=1, column=4 * num_not_same - 1, value='人工翻译')
            worksheet.cell(row=1, column=4 * num_not_same - 3, value=key)
            worksheet.cell(row=1, column=4 * num_not_same, value='校对')

            for i, value in enumerate(data2[key]['context']):

                worksheet.cell(row=i + 2, column=4 * num_not_same - 3, value=value)
                jifan = baidutrans(value)
                print(jifan)
                worksheet.cell(row=i + 2, column=4 * num_not_same - 2, value=jifan)
                try:
                    de = detect(jifan)
                except:
                    de = None
                if jifan is None or de != 'zh-cn':
                    continue
                jifan_list.append(jifan)

            for i, value in enumerate(data[key]['context']):

                index, distence = find_most_similar(value, jifan_list)
                if distence > 1:
                    worksheet.cell(row=index + 2, column=4 * num_not_same - 1, value=value)
                print(num_not_same, end='\r')
            jifan_list = ['这是一个空的文本']

workbook.save('naiya.xlsx')

print(num_same, num_not_same)
