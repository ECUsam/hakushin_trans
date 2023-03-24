import json
from openpyxl import Workbook
from langdetect import detect


# 创建一个Workbook对象
workbook = Workbook()

# 获取当前活动的Worksheet对象
worksheet = workbook.active

with open('data.json', 'r') as f:
    data = json.load(f)
with open('data2.json', 'r') as f:
    data2 = json.load(f)

data_test = data['po23_ed_C2']['context']
data_test2 = data2['po23_ed_C2']['context']

print(len(data_test))
print(len(data_test2))
print(data['po23_ed_C2']['context'])
print(data2['po23_ed_C2']['context'])

# 写入列表1
for i, value in enumerate(data_test, start=1):
    worksheet.cell(row=i, column=1, value=value)
    worksheet.cell(row=i, column=3, value=detect(value))
# 写入列表2
for i, value in enumerate(data_test2, start=1):
    worksheet.cell(row=i, column=2, value=value)


# 保存文件
workbook.save('output.xlsx')
