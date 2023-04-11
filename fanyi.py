import openpyxl
from openpyxl.utils import cell
import json

workbook = openpyxl.load_workbook('naiya.xlsx')
worksheet = workbook['Sheet']
num = 1


def read_text(column_number: int):
    res = []
    name = worksheet.cell(row=1, column=column_number).value
    column_letter = cell.get_column_letter(column_number + 1)
    for cel in worksheet[column_letter][1:]:
        if cel.value is not None:
            res.append(cel.value)
    return name, res


with open('naiya.json', 'r') as f:
    data = json.load(f)
with open('data2.json', 'r') as f:
    data2 = json.load(f)

while worksheet.cell(row=1, column=num).value is not None:
    name, res = read_text(num)
    assert name in data, name
    if len(data2[name]['context']) != len(res):
        print(data2[name]['context'])
        print(res)
        print(name)
        raise Exception(name)
    data[name]['context'] = res
    print(res)
    num += 3

with open('naiya.json', 'w') as f:
    json.dump(data, f, indent=4)
