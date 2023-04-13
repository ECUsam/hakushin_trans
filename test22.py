import json
from openpyxl import Workbook
# 创建一个Workbook对象
workbook = Workbook()

# 获取当前活动的Worksheet对象
worksheet = workbook.active

with open('data_.json', 'r', encoding='utf16') as f:
    data = json.load(f)
with open('data2.json', 'r', encoding='utf16') as f:
    data2 = json.load(f)

num_same = 0
num_not_same = 0

jifan_list = []

row_time = 3

for key in data2:
    if key in data:
        if len(data[key]['context']) == len(data2[key]['context']):
            num_same += 1
        else:
            num_not_same += 1
            worksheet.cell(row=1, column=row_time * num_not_same - 1, value='人工翻译')
            worksheet.cell(row=1, column=row_time * num_not_same - 2, value=key)
            worksheet.cell(row=1, column=row_time * num_not_same , value='校对')

            for i, value in enumerate(data2[key]['context']):
                worksheet.cell(row=i+2, column=row_time*num_not_same-2, value=value)
            for i, value in enumerate(data[key]['context']):
                print(value)
                worksheet.cell(row=i + 2, column=row_time * num_not_same - 1, value=value)
                # workbook.save('output.xlsx')

workbook.save('data_.xlsx')

print(num_same, num_not_same)
